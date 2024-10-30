import openpyxl
import pandas as pd
import yaml

with open('config.yaml', 'r') as yml:
  config = yaml.safe_load(yml)
LIST_FILE = config['population']['list']
FOLDER    = config['population']['folder']
ACTUAL1   = config['population']['actual1']
ACTUAL2   = config['population']['actual2']

def get_forcast():
  df = pd.read_csv(LIST_FILE, header=0)
  forcast_types = df.to_dict(orient='records')

  data = []
  for forcast_type in forcast_types:
    filename = FOLDER + forcast_type['file_name']
    workbook = openpyxl.load_workbook(filename)

    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
  
      nendo = sheet["A2"].value
      start = nendo.rfind('(')
      end = nendo.rfind('年')
      year = int(nendo[start+1:end-1])

      range = sheet["A5":"D59"] + sheet["F5":"I55"]
      for row in range:
        if year == 2020:
          continue
        data.append({
          'data_type' : forcast_type['category'],
          'year'         : year,
          'age'          : str(row[0].value),
          'total'        : row[1].value,
          'male'         : row[2].value,
          'female'       : row[3].value
        })

  return data

def get_history1():
  filename = FOLDER + ACTUAL1
  workbook = openpyxl.load_workbook(filename)
  data = []
  for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    if sheet_name == "大正9年～昭和15年":
      continue
    elif sheet_name == "昭和19年～29年":
      continue
    elif sheet_name == "昭和30年～46年":
      p = 1955
    elif sheet_name == "昭和47年～54年":
      p = 1972
    elif sheet_name == "昭和55年～平成12年":
      p = 1980

    rowdata = sheet["C13":"BM98"]
    data_by_col = pd.DataFrame(rowdata).T.values.tolist()
    data_by_year = []
    for i in range(0, len(data_by_col),3):
      data_by_year.append({
        'total'  : data_by_col[i],
        'male'   : data_by_col[i+1],
        'female' : data_by_col[i+2],
      })
    
    for year, data_by_age in enumerate(data_by_year):
      if data_by_age['total'][0].value is None:
        break
      for age in range(0, 86):
        if age == 85:
          str_age = '85+'
        else:
          str_age = str(age)
        data.append({
          'data_type' : 'history',
          'year' : p + year,
          'age' : str_age,
          'total'  : data_by_age['total'][age].value,
          'male'   : data_by_age['male'][age].value,
          'female' : data_by_age['female'][age].value
        })
    
  return data

def get_history2():
  filename = FOLDER + ACTUAL2
  workbook = openpyxl.load_workbook(filename)
  data = []
  for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    if sheet_name == "総人口（2000年～2015年）":
      p = 2001
    elif sheet_name == "日本人人口（2000年～2015年）":
      continue
    elif sheet_name == "総人口（2015年～2020年）":
      p = 2016
    elif sheet_name == "日本人人口（2015年～2020年）":
      continue

    rowdata = sheet["F14":"AX114"]
    data_by_col = pd.DataFrame(rowdata).T.values.tolist()
    data_by_year = []
    for i in range(0, len(data_by_col),3):
      data_by_year.append({
        'total'  : data_by_col[i],
        'male'   : data_by_col[i+1],
        'female' : data_by_col[i+2],
      })
    
    for year, data_by_age in enumerate(data_by_year):
      if data_by_age['total'][0].value is None:
        break
      if data_by_age['total'][90].value is None:
        max_age = 90
      else:
        max_age = 100

      for age in range(0, 101):

        if age >= max_age:
          str_age = str(max_age) + '+'
        else:
          str_age = str(age)

        if data_by_age['total'][age].value is None:
          continue

        data.append({
          'data_type' : 'history',
          'year' : p + year,
          'age' : str_age,
          'total'  : data_by_age['total'][age].value,
          'male'   : data_by_age['male'][age].value,
          'female' : data_by_age['female'][age].value
        })
    
  return data

if __name__ == '__main__':
  
  data = get_history1()
  data.extend(get_history2())
  data.extend(get_forcast())

  df = pd.DataFrame(data)
  df.to_excel('output\\population_master.xlsx', index=False, header=True)
